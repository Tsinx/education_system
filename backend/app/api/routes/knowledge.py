from collections import defaultdict
from dataclasses import dataclass
import io
from pathlib import Path
from typing import Iterable
import urllib.parse

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook

from app.schemas.material import DeleteKnowledgePointResponse
from app.services.material_pipeline import repository

router = APIRouter(prefix="/knowledge", tags=["knowledge"])
_TOPIC_TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "assets" / "importTopicTemplate.xlsx"


@dataclass(frozen=True)
class _ExportTopicRow:
    node_type: str
    level: int
    name: str
    description: str | None
    prerequisite_points: list[str]
    postrequisite_points: list[str]
    related_points: list[str]


def _normalize_topic_name(value: str | None) -> str:
    return (value or "").strip()


def _point_to_export_row(level: int, point) -> _ExportTopicRow:
    return _ExportTopicRow(
        node_type="知识点",
        level=max(1, min(7, level)),
        name=point.name,
        description=point.description or None,
        prerequisite_points=list(point.prerequisite_points),
        postrequisite_points=list(point.postrequisite_points),
        related_points=list(point.related_points),
    )


def _chapter_row(chapter_name: str, description: str | None = None) -> _ExportTopicRow:
    return _ExportTopicRow(
        node_type="分类",
        level=1,
        name=chapter_name,
        description=description or f"{chapter_name}章节知识组织节点",
        prerequisite_points=[],
        postrequisite_points=[],
        related_points=[],
    )


def _sort_points_for_export(points: list) -> list[_ExportTopicRow]:
    grouped_points: dict[tuple[str, int, str], list] = defaultdict(list)
    for point in points:
        grouped_points[(point.material_id, point.chapter_index, _normalize_topic_name(point.chapter_section))].append(point)

    export_rows: list[_ExportTopicRow] = []
    for group_key in sorted(grouped_points.keys(), key=lambda item: (item[0], item[1], item[2])):
        chapter_points = grouped_points[group_key]
        chapter_name = group_key[2]
        index_by_id = {point.id: idx for idx, point in enumerate(chapter_points)}
        children_by_parent: dict[str, list] = defaultdict(list)

        for point in chapter_points:
            children_by_parent[_normalize_topic_name(point.parent_name)].append(point)

        def order_children(candidates: Iterable, parent_point) -> list:
            ordered: list = []
            seen_ids: set[str] = set()
            child_names = [
                _normalize_topic_name(name)
                for name in getattr(parent_point, "child_points", [])
                if _normalize_topic_name(name)
            ]

            for child_name in child_names:
                for candidate in candidates:
                    if candidate.id in seen_ids:
                        continue
                    if _normalize_topic_name(candidate.name) == child_name:
                        ordered.append(candidate)
                        seen_ids.add(candidate.id)

            remaining = [candidate for candidate in candidates if candidate.id not in seen_ids]
            remaining.sort(key=lambda item: (index_by_id[item.id], _normalize_topic_name(item.name)))
            ordered.extend(remaining)
            return ordered

        chapter_root = next(
            (
                point
                for point in chapter_points
                if _normalize_topic_name(point.name) == chapter_name and not _normalize_topic_name(point.parent_name)
            ),
            None,
        )

        visited_ids: set[str] = set()

        def walk(nodes: list, level: int) -> None:
            for node in nodes:
                if node.id in visited_ids:
                    continue
                visited_ids.add(node.id)
                export_rows.append(_point_to_export_row(level, node))
                child_nodes = order_children(children_by_parent.get(_normalize_topic_name(node.name), []), node)
                walk(child_nodes, level + 1)

        if chapter_root is not None:
            visited_ids.add(chapter_root.id)
            export_rows.append(_chapter_row(chapter_root.name, chapter_root.description or None))
            walk(order_children(children_by_parent.get(_normalize_topic_name(chapter_root.name), []), chapter_root), 2)
        elif chapter_name:
            export_rows.append(_chapter_row(chapter_name))
            walk(order_children(children_by_parent.get("", []), None), 2)
        else:
            walk(order_children(children_by_parent.get("", []), None), 1)

        remaining_points = [point for point in chapter_points if point.id not in visited_ids]
        if remaining_points:
            remaining_points.sort(key=lambda item: (index_by_id[item.id], item.level, _normalize_topic_name(item.name)))
            walk(remaining_points, 2 if chapter_name else 1)

    return export_rows


@router.get("/summary")
def knowledge_summary() -> dict[str, object]:
    return {
        "project_id": "project_demo_001",
        "node_count": 28,
        "relation_count": 46,
        "top_topics": ["函数定义", "函数性质", "函数图像", "单调性"],
    }


@router.get("/courses/{course_id}/graph")
def course_knowledge_graph(course_id: str) -> dict[str, object]:
    try:
        course = repository.get_course(course_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="课程不存在") from None

    points = repository.list_course_knowledge_points(course_id)
    node_name_map: dict[str, list[str]] = defaultdict(list)
    child_edge_keys: set[tuple[str, str, str]] = set()
    relation_edge_keys: set[tuple[str, str, str]] = set()
    nodes: list[dict[str, object]] = []

    for point in points:
        node_id = point.id
        node_name_map[point.name].append(node_id)
        nodes.append(
            {
                "id": node_id,
                "label": point.name,
                "chapter": point.chapter_section,
                "chapter_index": point.chapter_index,
                "level": point.level,
                "parent_name": point.parent_name,
                "description": point.description,
                "prerequisite_points": point.prerequisite_points,
                "postrequisite_points": point.postrequisite_points,
                "related_points": point.related_points,
            }
        )

    edges: list[dict[str, object]] = []
    for point in points:
        if point.parent_name:
            for target_id in node_name_map.get(point.parent_name, []):
                edge_key = (target_id, point.id, "hierarchy")
                if edge_key in child_edge_keys:
                    continue
                child_edge_keys.add(edge_key)
                edges.append(
                    {
                        "id": f"{target_id}->{point.id}:hierarchy",
                        "source": target_id,
                        "target": point.id,
                        "type": "hierarchy",
                    }
                )

    relation_edges = repository.list_course_knowledge_edges(course_id)
    if relation_edges:
        for edge in relation_edges:
            source_id = str(edge.get("source_point_id", ""))
            target_id = str(edge.get("target_point_id", ""))
            relation_type = str(edge.get("relation_type", ""))
            if not source_id or not target_id or source_id == target_id:
                continue
            edge_key = (source_id, target_id, relation_type)
            if edge_key in relation_edge_keys:
                continue
            relation_edge_keys.add(edge_key)
            edges.append(
                {
                    "id": f"{source_id}->{target_id}:{relation_type}",
                    "source": source_id,
                    "target": target_id,
                    "type": relation_type,
                }
            )
    else:
        for point in points:
            for relation_type, names in (
                ("prerequisite", point.prerequisite_points),
                ("postrequisite", point.postrequisite_points),
                ("related", point.related_points),
            ):
                for name in names:
                    for target_id in node_name_map.get(name, []):
                        if target_id == point.id:
                            continue
                        if relation_type == "prerequisite":
                            source_id, target_node_id = target_id, point.id
                        else:
                            source_id, target_node_id = point.id, target_id
                        edge_key = (source_id, target_node_id, relation_type)
                        if edge_key in relation_edge_keys:
                            continue
                        relation_edge_keys.add(edge_key)
                        edges.append(
                            {
                                "id": f"{source_id}->{target_node_id}:{relation_type}",
                                "source": source_id,
                                "target": target_node_id,
                                "type": relation_type,
                            }
                        )

    chapters = sorted({point.chapter_section for point in points})
    return {
        "course_id": course_id,
        "course_name": course.name,
        "node_count": len(nodes),
        "edge_count": len(edges),
        "chapter_count": len(chapters),
        "chapters": chapters,
        "nodes": nodes,
        "edges": edges,
    }


@router.delete("/courses/{course_id}/nodes/{node_id}", response_model=DeleteKnowledgePointResponse)
def delete_course_knowledge_node(
    course_id: str,
    node_id: str,
    delete_descendants: bool = Query(default=False),
) -> DeleteKnowledgePointResponse:
    try:
        repository.get_course(course_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="课程不存在") from None

    try:
        result = repository.delete_course_knowledge_point(
            course_id=course_id,
            point_id=node_id,
            delete_descendants=delete_descendants,
        )
    except KeyError:
        raise HTTPException(status_code=404, detail="知识点不存在") from None

    message = (
        f"知识点已删除，并递归删除 {result['deleted_count'] - 1} 个子节点"
        if result["recursive"] and result["deleted_count"] > 1
        else f"知识点已删除，并提升 {result['promoted_count']} 个子节点"
        if result["promoted_count"] > 0
        else "知识点已删除"
    )
    return DeleteKnowledgePointResponse(
        message=message,
        deleted_count=int(result["deleted_count"]),
        promoted_count=int(result["promoted_count"]),
        recursive=bool(result["recursive"]),
    )


@router.get("/courses/{course_id}/graph-export")
def export_course_knowledge_graph(course_id: str) -> StreamingResponse:
    try:
        course = repository.get_course(course_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="课程不存在") from None

    if not _TOPIC_TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail="导出模板文件不存在")

    points = repository.list_course_knowledge_points(course_id)
    workbook = load_workbook(_TOPIC_TEMPLATE_PATH)
    sheet = workbook["Sheet3"] if "Sheet3" in workbook.sheetnames else workbook.worksheets[0]

    # Keep template instructions/header rows and rewrite rows from row 3.
    for row_idx in range(3, sheet.max_row + 1):
        for col_idx in range(1, 15):
            sheet.cell(row=row_idx, column=col_idx).value = None

    current_row = 3
    for row in _sort_points_for_export(points):
        # New template uses column A for node type, and level columns start from column B.
        sheet.cell(row=current_row, column=1).value = row.node_type
        sheet.cell(row=current_row, column=row.level + 1).value = row.name
        sheet.cell(row=current_row, column=9).value = ";".join(row.prerequisite_points) or None
        sheet.cell(row=current_row, column=10).value = ";".join(row.postrequisite_points) or None
        sheet.cell(row=current_row, column=11).value = ";".join(row.related_points) or None
        sheet.cell(row=current_row, column=14).value = row.description or None
        current_row += 1

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    safe_course_name = "".join(ch for ch in course.name if ch not in '\\/:*?"<>|').strip() or course_id
    filename = urllib.parse.quote(f"{safe_course_name}_学习通知识图谱导入.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
